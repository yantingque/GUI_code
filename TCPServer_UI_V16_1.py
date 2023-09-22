import socket
import threading

from PyQt5.QtCore import QTimer

from openpyxl import Workbook
import time


from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtWidgets import (QApplication, QMessageBox)
from mplwidget import MplWidget

import pandas as pd
from scipy.signal import butter, filtfilt

import numpy as np


HOST = '0.0.0.0'
PORT = 8000
client_ip = '0.0.0.0'

time_reset = 500

s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
s.connect(("8.8.8.8", 80))
HOST = s.getsockname()[0]
s.close()

t_list = []
accelerometer_x = []    #加速度計
accelerometer_y = []
accelerometer_z = []


excel_state = False
accelerometer_excel = []


t=0
origin = 0
lock = 0
reset_num = 0
ex_row = 1
click_state = False
msg_error = False
LC = 0
HC = 0

sampleRate_rate = 1250

class WorkerThread(QThread): 
    trigger = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.host = HOST
        self.port = PORT
        
    
    def run(self):
        self.doConnect()
        while True:
            try:
                client, addr = self.server.accept()
                # self.recv_msg(client, addr)
                #threading.Thread(target=self.send_msg, args=(client, addr)).start()
                threading.Thread(target=self.recv_msg, args=(client, addr)).start()
            except socket.error:
                print('socket connect error, doing connect host/port:{}/{}'.format(self.host, self.port))
                self.doConnect()
            except Exception as e:
                print('other error occur:{}'.format(e))
    
    def doConnect(self):
        try:
            self.server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            # 防止socket server重启后端口被占用（socket.error: [Errno 98] Address already in use）
            self.server.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            self.server.bind((self.host, self.port))
            self.server.listen(5)
            print('-----------------------------------------------------------')
            print("esWsServer host:{}/port:{} started listen...".format(self.host, self.port))
            print('-----------------------------------------------------------')
        except Exception as e:
            print('start ws server error:{}'.format(str(e)))
    
    def recv_msg(self, client, addr):
        global client_ip
        try:
            print('Accept new connection from {0}'.format(addr))
            client_ip = addr[0]
            print('Client IP: ',client_ip)
            while 1:
                data = str(client.recv(2048), encoding='utf-8', errors='ignore')
                if data == "IMU ERR\n":
                    print('Client data is:', data)
                elif len(data)>=1:
                    self.trigger.emit(data)

        except Exception as e:
            print('recv_msg:{}'.format(e))

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1260, 970)
        MainWindow.setMinimumSize(QtCore.QSize(1260, 970))
        MainWindow.setMaximumSize(QtCore.QSize(1260, 970))
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.horizontalLayoutWidget_4 = QtWidgets.QWidget(self.centralwidget)
        self.horizontalLayoutWidget_4.setGeometry(QtCore.QRect(-1, 0, 1251, 921))
        self.horizontalLayoutWidget_4.setObjectName("horizontalLayoutWidget_4")
        self.horizontalLayout_19 = QtWidgets.QHBoxLayout(self.horizontalLayoutWidget_4)
        self.horizontalLayout_19.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_19.setObjectName("horizontalLayout_19")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.groupBox = QtWidgets.QGroupBox(self.horizontalLayoutWidget_4)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.groupBox.setFont(font)
        self.groupBox.setObjectName("groupBox")
        self.verticalLayoutWidget_5 = QtWidgets.QWidget(self.groupBox)
        self.verticalLayoutWidget_5.setGeometry(QtCore.QRect(10, 20, 241, 71))
        self.verticalLayoutWidget_5.setObjectName("verticalLayoutWidget_5")
        self.verticalLayout_5 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_5)
        self.verticalLayout_5.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_5.setObjectName("verticalLayout_5")
        self.verticalLayout_6 = QtWidgets.QVBoxLayout()
        self.verticalLayout_6.setObjectName("verticalLayout_6")
        self.line_IP_s = QtWidgets.QLineEdit(self.verticalLayoutWidget_5)
        self.line_IP_s.setObjectName("line_IP_s")
        self.verticalLayout_6.addWidget(self.line_IP_s)
        self.verticalLayout_5.addLayout(self.verticalLayout_6)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        spacerItem = QtWidgets.QSpacerItem(50, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem)
        self.btn_IP = QtWidgets.QPushButton(self.verticalLayoutWidget_5)
        self.btn_IP.setObjectName("btn_IP")
        self.horizontalLayout.addWidget(self.btn_IP)
        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem1)
        self.horizontalLayout.setStretch(0, 20)
        self.horizontalLayout.setStretch(1, 3)
        self.horizontalLayout.setStretch(2, 1)
        self.verticalLayout_5.addLayout(self.horizontalLayout)
        self.verticalLayout_5.setStretch(0, 5)
        self.verticalLayout_5.setStretch(1, 5)
        self.verticalLayout.addWidget(self.groupBox)
        self.groupBox_4 = QtWidgets.QGroupBox(self.horizontalLayoutWidget_4)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.groupBox_4.setFont(font)
        self.groupBox_4.setObjectName("groupBox_4")
        self.line_IP_c = QtWidgets.QLineEdit(self.groupBox_4)
        self.line_IP_c.setGeometry(QtCore.QRect(10, 20, 237, 29))
        self.line_IP_c.setObjectName("line_IP_c")
        self.verticalLayout.addWidget(self.groupBox_4)
        self.groupBox_2 = QtWidgets.QGroupBox(self.horizontalLayoutWidget_4)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.groupBox_2.setFont(font)
        self.groupBox_2.setObjectName("groupBox_2")
        self.verticalLayoutWidget_2 = QtWidgets.QWidget(self.groupBox_2)
        self.verticalLayoutWidget_2.setGeometry(QtCore.QRect(10, 20, 221, 161))
        self.verticalLayoutWidget_2.setObjectName("verticalLayoutWidget_2")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_2)
        self.verticalLayout_2.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setSpacing(1)
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.ckb_acc = QtWidgets.QCheckBox(self.verticalLayoutWidget_2)
        self.ckb_acc.setObjectName("ckb_acc")
        self.horizontalLayout_2.addWidget(self.ckb_acc)
        self.ckb_acc_out = QtWidgets.QCheckBox(self.verticalLayoutWidget_2)
        self.ckb_acc_out.setObjectName("ckb_acc_out")
        self.horizontalLayout_2.addWidget(self.ckb_acc_out)
        self.verticalLayout_2.addLayout(self.horizontalLayout_2)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        spacerItem2 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_3.addItem(spacerItem2)
        self.verticalLayout_3 = QtWidgets.QVBoxLayout()
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.ckb_X = QtWidgets.QCheckBox(self.verticalLayoutWidget_2)
        self.ckb_X.setObjectName("ckb_X")
        self.verticalLayout_3.addWidget(self.ckb_X)
        self.ckb_Y = QtWidgets.QCheckBox(self.verticalLayoutWidget_2)
        self.ckb_Y.setObjectName("ckb_Y")
        self.verticalLayout_3.addWidget(self.ckb_Y)
        self.ckb_Z = QtWidgets.QCheckBox(self.verticalLayoutWidget_2)
        self.ckb_Z.setObjectName("ckb_Z")
        self.verticalLayout_3.addWidget(self.ckb_Z)
        self.horizontalLayout_3.addLayout(self.verticalLayout_3)
        self.horizontalLayout_3.setStretch(0, 1)
        self.horizontalLayout_3.setStretch(1, 9)
        self.verticalLayout_2.addLayout(self.horizontalLayout_3)
        self.verticalLayout_2.setStretch(0, 2)
        self.verticalLayout_2.setStretch(1, 6)
        self.verticalLayout.addWidget(self.groupBox_2)
        self.groupBox_3 = QtWidgets.QGroupBox(self.horizontalLayoutWidget_4)
        self.groupBox_3.setObjectName("groupBox_3")
        self.layoutWidget = QtWidgets.QWidget(self.groupBox_3)
        self.layoutWidget.setGeometry(QtCore.QRect(10, 20, 229, 171))
        self.layoutWidget.setObjectName("layoutWidget")
        self.verticalLayout_7 = QtWidgets.QVBoxLayout(self.layoutWidget)
        self.verticalLayout_7.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_7.setObjectName("verticalLayout_7")
        self.horizontalLayout_8 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_8.setObjectName("horizontalLayout_8")
        spacerItem3 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_8.addItem(spacerItem3)
        self.label_5 = QtWidgets.QLabel(self.layoutWidget)
        self.label_5.setObjectName("label_5")
        self.horizontalLayout_8.addWidget(self.label_5)
        self.line_accX = QtWidgets.QLineEdit(self.layoutWidget)
        self.line_accX.setObjectName("line_accX")
        self.horizontalLayout_8.addWidget(self.line_accX)
        self.horizontalLayout_8.setStretch(0, 1)
        self.horizontalLayout_8.setStretch(1, 1)
        self.horizontalLayout_8.setStretch(2, 7)
        self.verticalLayout_7.addLayout(self.horizontalLayout_8)
        self.horizontalLayout_7 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_7.setObjectName("horizontalLayout_7")
        spacerItem4 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_7.addItem(spacerItem4)
        self.label_4 = QtWidgets.QLabel(self.layoutWidget)
        self.label_4.setObjectName("label_4")
        self.horizontalLayout_7.addWidget(self.label_4)
        self.line_accY = QtWidgets.QLineEdit(self.layoutWidget)
        self.line_accY.setObjectName("line_accY")
        self.horizontalLayout_7.addWidget(self.line_accY)
        self.horizontalLayout_7.setStretch(0, 1)
        self.horizontalLayout_7.setStretch(1, 1)
        self.horizontalLayout_7.setStretch(2, 7)
        self.verticalLayout_7.addLayout(self.horizontalLayout_7)
        self.horizontalLayout_9 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_9.setObjectName("horizontalLayout_9")
        spacerItem5 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_9.addItem(spacerItem5)
        self.label_6 = QtWidgets.QLabel(self.layoutWidget)
        self.label_6.setObjectName("label_6")
        self.horizontalLayout_9.addWidget(self.label_6)
        self.line_accZ = QtWidgets.QLineEdit(self.layoutWidget)
        self.line_accZ.setObjectName("line_accZ")
        self.horizontalLayout_9.addWidget(self.line_accZ)
        self.horizontalLayout_9.setStretch(0, 1)
        self.horizontalLayout_9.setStretch(1, 1)
        self.horizontalLayout_9.setStretch(2, 7)
        self.verticalLayout_7.addLayout(self.horizontalLayout_9)
        self.verticalLayout_7.setStretch(0, 2)
        self.verticalLayout_7.setStretch(1, 2)
        self.verticalLayout_7.setStretch(2, 2)
        self.verticalLayout.addWidget(self.groupBox_3)
        self.groupBox_6 = QtWidgets.QGroupBox(self.horizontalLayoutWidget_4)
        self.groupBox_6.setObjectName("groupBox_6")
        self.layoutWidget_2 = QtWidgets.QWidget(self.groupBox_6)
        self.layoutWidget_2.setGeometry(QtCore.QRect(10, 20, 229, 171))
        self.layoutWidget_2.setObjectName("layoutWidget_2")
        self.verticalLayout_8 = QtWidgets.QVBoxLayout(self.layoutWidget_2)
        self.verticalLayout_8.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_8.setObjectName("verticalLayout_8")
        self.horizontalLayout_10 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_10.setObjectName("horizontalLayout_10")
        spacerItem6 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_10.addItem(spacerItem6)
        self.label_7 = QtWidgets.QLabel(self.layoutWidget_2)
        self.label_7.setObjectName("label_7")
        self.horizontalLayout_10.addWidget(self.label_7)
        self.line_low = QtWidgets.QLineEdit(self.layoutWidget_2)
        self.line_low.setObjectName("line_low")
        self.horizontalLayout_10.addWidget(self.line_low)
        self.horizontalLayout_10.setStretch(0, 1)
        self.horizontalLayout_10.setStretch(1, 1)
        self.horizontalLayout_10.setStretch(2, 7)
        self.verticalLayout_8.addLayout(self.horizontalLayout_10)
        self.horizontalLayout_11 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_11.setObjectName("horizontalLayout_11")
        spacerItem7 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_11.addItem(spacerItem7)
        self.label_8 = QtWidgets.QLabel(self.layoutWidget_2)
        self.label_8.setObjectName("label_8")
        self.horizontalLayout_11.addWidget(self.label_8)
        self.line_high = QtWidgets.QLineEdit(self.layoutWidget_2)
        self.line_high.setObjectName("line_high")
        self.horizontalLayout_11.addWidget(self.line_high)
        self.horizontalLayout_11.setStretch(0, 1)
        self.horizontalLayout_11.setStretch(1, 1)
        self.horizontalLayout_11.setStretch(2, 7)
        self.verticalLayout_8.addLayout(self.horizontalLayout_11)
        self.verticalLayout_8.setStretch(0, 2)
        self.verticalLayout_8.setStretch(1, 2)
        self.verticalLayout.addWidget(self.groupBox_6)
        self.verticalLayout.setStretch(0, 1)
        self.verticalLayout.setStretch(1, 1)
        self.verticalLayout.setStretch(2, 2)
        self.verticalLayout.setStretch(3, 2)
        self.verticalLayout.setStretch(4, 2)
        self.horizontalLayout_19.addLayout(self.verticalLayout)
        self.groupBox_5 = QtWidgets.QGroupBox(self.horizontalLayoutWidget_4)
        font = QtGui.QFont()
        font.setFamily("Adobe Devanagari")
        font.setPointSize(10)
        self.groupBox_5.setFont(font)
        self.groupBox_5.setObjectName("groupBox_5")
        self.verticalLayoutWidget_4 = QtWidgets.QWidget(self.groupBox_5)
        self.verticalLayoutWidget_4.setGeometry(QtCore.QRect(10, 20, 981, 891))
        self.verticalLayoutWidget_4.setObjectName("verticalLayoutWidget_4")
        self.verticalLayout_13 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_4)
        self.verticalLayout_13.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_13.setObjectName("verticalLayout_13")
        self.MplWidget = MplWidget(self.verticalLayoutWidget_4)
        self.MplWidget.setObjectName("MplWidget")
        self.verticalLayout_13.addWidget(self.MplWidget)
        self.horizontalLayout_18 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_18.setObjectName("horizontalLayout_18")
        self.btn_record = QtWidgets.QPushButton(self.verticalLayoutWidget_4)
        self.btn_record.setObjectName("btn_record")
        self.horizontalLayout_18.addWidget(self.btn_record)
        self.btn_save = QtWidgets.QPushButton(self.verticalLayoutWidget_4)
        self.btn_save.setObjectName("btn_save")
        self.horizontalLayout_18.addWidget(self.btn_save)
        spacerItem8 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_18.addItem(spacerItem8)
        self.btn_stop = QtWidgets.QPushButton(self.verticalLayoutWidget_4)
        self.btn_stop.setObjectName("btn_stop")
        self.horizontalLayout_18.addWidget(self.btn_stop)
        self.verticalLayout_13.addLayout(self.horizontalLayout_18)
        self.verticalLayout_13.setStretch(0, 18)
        self.verticalLayout_13.setStretch(1, 1)
        self.horizontalLayout_19.addWidget(self.groupBox_5)
        self.horizontalLayout_19.setStretch(0, 2)
        self.horizontalLayout_19.setStretch(1, 8)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1260, 25))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        #設定
        #------------------------元件初始化--------------------------#
        self.line_IP_c.setReadOnly(True)
        self.line_IP_s.setReadOnly(True)
        self.line_accX.setReadOnly(True)
        self.line_accY.setReadOnly(True)
        self.line_accZ.setReadOnly(True)
        self.btn_save.setEnabled(False)
        self.line_low.setText('250')
        self.line_high.setText('499')
        #------------------------功能初始化-------------------------------#
        self.ckb_acc.setChecked(False)
        self.ckb_X.setChecked(False)
        self.ckb_Y.setChecked(False)
        self.ckb_Z.setChecked(False)
        
        #--------------------------主功能---------------------------------#
        self.btn_stop.clicked.connect(self.btn_lock)#圖表暫停功能
        self.ckb_acc.clicked.connect(self.click_acc)
        self.ckb_acc_out.clicked.connect(self.click_acc_out)
        self.ckb_X.clicked.connect(self.click_X)
        self.ckb_Y.clicked.connect(self.click_Y)
        self.ckb_Z.clicked.connect(self.click_Z)
        self.btn_record.clicked.connect(self.click_record)
        self.btn_save.clicked.connect(self.save_excel)

        #線程
        self.work = WorkerThread()#呼叫TCP線程
        self.btn_IP.clicked.connect(self.startThread)

        #Timer(更新曲線圖)
        self.drawtimer = QtCore.QTimer()#Timer
        self.drawtimer.timeout.connect(self.update_graph)
        self.drawtimer.start(5)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        
    def startThread(self):
        self.btn_IP.setEnabled(False)
        self.line_IP_s.setText(HOST)
        self.work.start()
        self.work.trigger.connect(self.update_data)
    
    def update_data(self, msg):
        global client_ip
        global t
        global origin
        global t_list
        global accelerometer_x
        global accelerometer_y
        global accelerometer_z
        global msg_error
        
        self.line_IP_c.setText(client_ip)
        list_msg = msg.split('#')
        
        if(len(list_msg)>0):
            while len(list_msg)>0:
                list_data = list_msg[0].split(',')
                if len(list_data)==5 or len(list_data)==6:
                    list_data.pop(0)

                if len(list_data)==6:
                    print(list_data)

                list_msg.pop(0)
                msg_error=False
                
                try:
                    int(list_data[0])
                    int(list_data[1])
                    int(list_data[2])
                except:
                    msg_error=True

                if(msg_error==False):
                    if len(list_data)==4:
                        if(excel_state == True):                                #判斷是否需要紀錄  
                            accelerometer_data = []
                            accelerometer_data.append(round(float(list_data[0])*0.000488,4))
                            accelerometer_data.append(round(float(list_data[1])*0.000488,4))
                            accelerometer_data.append(round(float(list_data[2])*0.000488,4))                             
                            accelerometer_excel.append(accelerometer_data)        

                        self.line_accX.setText(str(round(float(list_data[0])*0.000488,4)))
                        accelerometer_x.append(float(list_data.pop(0))*0.000488) #拿取資料後刪除該資料
                        self.line_accY.setText(str(round(float(list_data[0])*0.000488,4)))
                        accelerometer_y.append(float(list_data.pop(0))*0.000488)
                        self.line_accZ.setText(str(round(float(list_data[0])*0.000488,4)))
                        accelerometer_z.append(float(list_data.pop(0))*0.000488)

                        list_data.pop(0)
                        t = t + 1
                        #t_list.append(t)
                        t_list.append(t/sampleRate_rate)
                        if t > sampleRate_rate:
                            origin = t-sampleRate_rate
                        if t>sampleRate_rate:
                            t_list.pop(0)
                            accelerometer_x.pop(0)
                            accelerometer_y.pop(0)
                            accelerometer_z.pop(0)
                    elif len(list_data) == 5:
                        print(list_data)


    def update_graph(self):
        global origin
        global t_list
        global accelerometer_x
        global accelerometer_y
        global accelerometer_z
        global LC
        global HC
        listx_max = 0.5
        listy_max = 0.5
        listz_max = 0.5
        listx_min = -0.5
        listy_min = -0.5
        listz_min = -0.5
        # 設置濾波器參數
        highcut = HC  # 高頻截止頻率，單位為Hz
        fs = 1100  # 采樣頻率，單位為Hz

        if lock == 0:
            if len(accelerometer_x)>0:
                listx_max = max(accelerometer_x) + 0.2
                listy_max = max(accelerometer_y) + 0.2
                listz_max = max(accelerometer_z) + 0.2
                listx_min = min(accelerometer_x) - 0.2
                listy_min = min(accelerometer_y) - 0.2
                listz_min = min(accelerometer_z) - 0.2
            if self.ckb_acc.isChecked():
                if click_state == True:
                    self.MplWidget.canvas.figure.clf()
                    self.MplWidget.canvas.axes0 = self.MplWidget.canvas.figure.add_subplot(311)
                    self.MplWidget.canvas.axes1 = self.MplWidget.canvas.figure.add_subplot(312)
                    self.MplWidget.canvas.axes2 = self.MplWidget.canvas.figure.add_subplot(313)

                    self.MplWidget.canvas.axes0.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes0.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes0.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    self.MplWidget.canvas.axes0.set_ylim([listx_min, listx_max])
                    self.MplWidget.canvas.axes0.plot(t_list, accelerometer_x, color='red') 

                    self.MplWidget.canvas.axes1.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes1.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes1.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    self.MplWidget.canvas.axes1.set_ylim([listy_min, listy_max])
                    self.MplWidget.canvas.axes1.plot(t_list, accelerometer_y, color='green') 

                    self.MplWidget.canvas.axes2.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes2.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes2.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    self.MplWidget.canvas.axes2.set_ylim([listz_min, listz_max])
                    self.MplWidget.canvas.axes2.plot(t_list, accelerometer_z, color='blue')

                    self.MplWidget.canvas.axes0.set_title('Accelerometer (X)') 
                    self.MplWidget.canvas.axes1.set_title('Accelerometer (Y)') 
                    self.MplWidget.canvas.axes2.set_title('Accelerometer (Z)') 
                    self.MplWidget.canvas.figure.tight_layout() #隔開兩個圖
                    self.MplWidget.canvas.draw()

            elif self.ckb_acc_out.isChecked():
                #四分位距法去除離群值
                out_pd = pd.Index(accelerometer_x)
                out_x = remove_outliers(out_pd)

                out_pd = pd.Index(accelerometer_y)
                out_y = remove_outliers(out_pd)

                out_pd = pd.Index(accelerometer_z)
                out_z = remove_outliers(out_pd)
                
                # 設計低通濾波器
                cutoff = highcut  # 截止頻率
                nyq = 0.5 * fs  # Nyquist頻率
                normal_cutoff = cutoff / nyq
                b, a = butter(4, normal_cutoff, btype='low', analog=False)

                df_x = filtfilt(b, a, out_x)
                df_y = filtfilt(b, a, out_y)
                df_z = filtfilt(b, a, out_z)

                
                if click_state == True:
                    self.MplWidget.canvas.figure.clf()
                    self.MplWidget.canvas.axes0 = self.MplWidget.canvas.figure.add_subplot(311)
                    self.MplWidget.canvas.axes1 = self.MplWidget.canvas.figure.add_subplot(312)
                    self.MplWidget.canvas.axes2 = self.MplWidget.canvas.figure.add_subplot(313)

                    self.MplWidget.canvas.axes0.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes0.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes0.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    self.MplWidget.canvas.axes0.set_ylim([listx_min, listx_max])
                    self.MplWidget.canvas.axes0.plot(t_list, df_x, color='red') 

                    self.MplWidget.canvas.axes1.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes1.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes1.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    self.MplWidget.canvas.axes1.set_ylim([listy_min, listy_max])
                    self.MplWidget.canvas.axes1.plot(t_list, df_y, color='green') 

                    self.MplWidget.canvas.axes2.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes2.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes2.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    self.MplWidget.canvas.axes2.set_ylim([listz_min, listz_max])
                    self.MplWidget.canvas.axes2.plot(t_list, df_z, color='blue')

                    self.MplWidget.canvas.axes0.set_title('Accelerometer (X)') 
                    self.MplWidget.canvas.axes1.set_title('Accelerometer (Y)') 
                    self.MplWidget.canvas.axes2.set_title('Accelerometer (Z)') 
                    self.MplWidget.canvas.figure.tight_layout() #隔開兩個圖
                    self.MplWidget.canvas.draw()

            elif self.ckb_X.isChecked() and self.ckb_Y.isChecked() and self.ckb_Z.isChecked():
                if click_state == True:
                    self.MplWidget.canvas.figure.clf()
                    self.MplWidget.canvas.axes0 = self.MplWidget.canvas.figure.add_subplot(311)
                    self.MplWidget.canvas.axes1 = self.MplWidget.canvas.figure.add_subplot(312)
                    self.MplWidget.canvas.axes2 = self.MplWidget.canvas.figure.add_subplot(313)

                    self.MplWidget.canvas.axes0.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes0.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes0.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    self.MplWidget.canvas.axes0.set_ylim([listx_min, listx_max])
                    self.MplWidget.canvas.axes0.plot(t_list, accelerometer_x, color='red') 

                    self.MplWidget.canvas.axes1.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes1.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes1.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    #self.MplWidget.canvas.axes1.set_xlim([origin, t])
                    self.MplWidget.canvas.axes1.set_ylim([listy_min, listy_max])
                    self.MplWidget.canvas.axes1.plot(t_list, accelerometer_y, color='green') 

                    self.MplWidget.canvas.axes2.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes2.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes2.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    #self.MplWidget.canvas.axes2.set_xlim([origin, t])
                    self.MplWidget.canvas.axes2.set_ylim([listz_min, listz_max])
                    self.MplWidget.canvas.axes2.plot(t_list, accelerometer_z, color='blue')

                    self.MplWidget.canvas.axes0.set_title('Accelerometer (X)') 
                    self.MplWidget.canvas.axes1.set_title('Accelerometer (Y)') 
                    self.MplWidget.canvas.axes2.set_title('Accelerometer (Z)') 
                    self.MplWidget.canvas.figure.tight_layout() #隔開兩個圖
                    self.MplWidget.canvas.draw()
        
            elif self.ckb_X.isChecked() and self.ckb_Y.isChecked():
                if click_state == True:
                    self.MplWidget.canvas.figure.clf()
                    self.MplWidget.canvas.axes0 = self.MplWidget.canvas.figure.add_subplot(211)
                    self.MplWidget.canvas.axes1 = self.MplWidget.canvas.figure.add_subplot(212)

                    self.MplWidget.canvas.axes0.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes0.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes0.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    self.MplWidget.canvas.axes0.set_ylim([listx_min, listx_max])
                    self.MplWidget.canvas.axes0.plot(t_list, accelerometer_x, color='red') 

                    self.MplWidget.canvas.axes1.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes1.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes1.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    self.MplWidget.canvas.axes1.set_ylim([listy_min, listy_max])
                    self.MplWidget.canvas.axes1.plot(t_list, accelerometer_y, color='green') 

                    self.MplWidget.canvas.axes0.set_title('Accelerometer (X)') 
                    self.MplWidget.canvas.axes1.set_title('Accelerometer (Y)')  
                    self.MplWidget.canvas.figure.tight_layout() #隔開兩個圖
                    self.MplWidget.canvas.draw()
                
            elif self.ckb_X.isChecked() and self.ckb_Z.isChecked():
                if click_state == True:
                    self.MplWidget.canvas.figure.clf()
                    self.MplWidget.canvas.axes0 = self.MplWidget.canvas.figure.add_subplot(211)
                    self.MplWidget.canvas.axes1 = self.MplWidget.canvas.figure.add_subplot(212)

                    self.MplWidget.canvas.axes0.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes0.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes0.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    self.MplWidget.canvas.axes0.set_ylim([listx_min, listx_max])
                    self.MplWidget.canvas.axes0.plot(t_list, accelerometer_x, color='red') 

                    self.MplWidget.canvas.axes1.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes1.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes1.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    self.MplWidget.canvas.axes1.set_ylim([listz_min, listz_max])
                    self.MplWidget.canvas.axes1.plot(t_list, accelerometer_z, color='green') 

                    self.MplWidget.canvas.axes0.set_title('Accelerometer (X)') 
                    self.MplWidget.canvas.axes1.set_title('Accelerometer (Z)')  
                    self.MplWidget.canvas.figure.tight_layout() #隔開兩個圖
                    self.MplWidget.canvas.draw()

            elif self.ckb_Y.isChecked() and self.ckb_Z.isChecked():
                if click_state == True:
                    self.MplWidget.canvas.figure.clf()
                    self.MplWidget.canvas.axes0 = self.MplWidget.canvas.figure.add_subplot(211)
                    self.MplWidget.canvas.axes1 = self.MplWidget.canvas.figure.add_subplot(212)

                    self.MplWidget.canvas.axes0.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes0.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes0.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    self.MplWidget.canvas.axes0.set_ylim([listy_min, listy_max])
                    self.MplWidget.canvas.axes0.plot(t_list, accelerometer_y, color='red') 

                    self.MplWidget.canvas.axes1.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes1.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes1.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    self.MplWidget.canvas.axes1.set_ylim([listz_min, listz_max])
                    self.MplWidget.canvas.axes1.plot(t_list, accelerometer_z, color='green') 

                    self.MplWidget.canvas.axes0.set_title('Accelerometer (Y)') 
                    self.MplWidget.canvas.axes1.set_title('Accelerometer (Z)')  
                    self.MplWidget.canvas.figure.tight_layout() #隔開兩個圖
                    self.MplWidget.canvas.draw()
                
            elif self.ckb_X.isChecked():
                
                #四分位距法去除離群值
                out_pd = pd.Index(accelerometer_x)
                out_x = remove_outliers(out_pd)

                # 計算濾波器係數
                '''
                nyquist_freq = 0.5 * fs
                low = lowcut / nyquist_freq
                high = highcut / nyquist_freq
                b, a = butter(order, [low, high], btype="bandstop")
                df_x = filtfilt(b, a, out_x)
                '''
                # 設計低通濾波器
                cutoff = highcut  # 截止頻率
                nyq = 0.5 * fs  # Nyquist頻率
                normal_cutoff = cutoff / nyq
                b, a = butter(4, normal_cutoff, btype='low', analog=False)

                # 使用濾波器濾波信號
                filtered_x = filtfilt(b, a, accelerometer_x)

                # 設置參數
                N = sampleRate_rate      # 訊號長度
                T = 1 / sampleRate_rate   # 取樣週期
                y_f = np.fft.fft(filtered_x)
                x_f = np.linspace(0.0, 1.0/(2.0*T), N//2)

                if click_state == True:
                    self.MplWidget.canvas.figure.clf()
                    self.MplWidget.canvas.axes0 = self.MplWidget.canvas.figure.add_subplot(411)
                    self.MplWidget.canvas.axes1 = self.MplWidget.canvas.figure.add_subplot(412)
                    self.MplWidget.canvas.axes2 = self.MplWidget.canvas.figure.add_subplot(413)
                    self.MplWidget.canvas.axes3 = self.MplWidget.canvas.figure.add_subplot(414)

                    self.MplWidget.canvas.axes0.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes0.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes0.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    #self.MplWidget.canvas.axes0.set_xlim([origin, t])
                    self.MplWidget.canvas.axes0.set_ylim([listx_min, listx_max])
                    self.MplWidget.canvas.axes0.plot(t_list, accelerometer_x, color='red')  
                    self.MplWidget.canvas.axes0.set_title('Accelerometer (X)')

                    self.MplWidget.canvas.axes1.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes1.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes1.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    #self.MplWidget.canvas.axes1.set_xlim([origin, t])
                    self.MplWidget.canvas.axes1.set_ylim([listx_min, listx_max])
                    self.MplWidget.canvas.axes1.plot(t_list, out_x, color='#FF8328')  
                    self.MplWidget.canvas.axes1.set_title('Remove Outliers Accelerometer (X)')

                    self.MplWidget.canvas.axes2.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes2.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes2.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    #self.MplWidget.canvas.axes2.set_xlim([origin, t])
                    self.MplWidget.canvas.axes2.set_ylim([listx_min, listx_max])
                    self.MplWidget.canvas.axes2.plot(t_list, filtered_x, color='blue')  
                    self.MplWidget.canvas.axes2.set_title('Filt Accelerometer (X)') 

                    self.MplWidget.canvas.axes3.set_xlabel("Frequency(Hz)")
                    self.MplWidget.canvas.axes3.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes3.set_xlim([0, 200])
                    self.MplWidget.canvas.axes3.set_ylim([0, 1])
                    self.MplWidget.canvas.axes3.plot(x_f, 2.0/N * np.abs(y_f[:N//2]), color='green')  
                    self.MplWidget.canvas.axes3.set_title('FFT Accelerometer (X)')     

                    self.MplWidget.canvas.figure.tight_layout() #隔開兩個圖
                    self.MplWidget.canvas.draw()

            elif self.ckb_Y.isChecked():
                #四分位距法去除離群值
                out_pd = pd.Index(accelerometer_y)
                out_y = remove_outliers(out_pd)

                '''
                nyquist_freq = 0.5 * fs
                low = lowcut / nyquist_freq
                high = highcut / nyquist_freq
                b, a = butter(order, [low, high], btype="bandstop")
                df_x = filtfilt(b, a, out_x)
                '''
                # 設計低通濾波器
                cutoff = highcut  # 截止頻率
                nyq = 0.5 * fs  # Nyquist頻率
                normal_cutoff = cutoff / nyq
                b, a = butter(4, normal_cutoff, btype='low', analog=False)

                # 使用濾波器濾波信號
                filtered_y = filtfilt(b, a, accelerometer_y)

                # 設置參數
                N = sampleRate_rate      # 訊號長度
                T = 1 / sampleRate_rate   # 取樣週期
                y_f = np.fft.fft(filtered_y)
                x_f = np.linspace(0.0, 1.0/(2.0*T), N//2)

                if click_state == True:
                    self.MplWidget.canvas.figure.clf()
                    self.MplWidget.canvas.axes0 = self.MplWidget.canvas.figure.add_subplot(411)
                    self.MplWidget.canvas.axes1 = self.MplWidget.canvas.figure.add_subplot(412)
                    self.MplWidget.canvas.axes2 = self.MplWidget.canvas.figure.add_subplot(413)
                    self.MplWidget.canvas.axes3 = self.MplWidget.canvas.figure.add_subplot(414)

                    self.MplWidget.canvas.axes0.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes0.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes0.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    #self.MplWidget.canvas.axes0.set_xlim([origin, t])
                    self.MplWidget.canvas.axes0.set_ylim([listy_min, listy_max])
                    self.MplWidget.canvas.axes0.plot(t_list, accelerometer_y, color='red')  
                    self.MplWidget.canvas.axes0.set_title('Accelerometer (Y)')

                    self.MplWidget.canvas.axes1.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes1.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes1.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    #self.MplWidget.canvas.axes1.set_xlim([origin, t])
                    self.MplWidget.canvas.axes1.set_ylim([listy_min, listy_max])
                    self.MplWidget.canvas.axes1.plot(t_list, out_y, color='#FF8328')  
                    self.MplWidget.canvas.axes1.set_title('Remove Outliers Accelerometer (Y)')

                    self.MplWidget.canvas.axes2.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes2.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes2.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    #self.MplWidget.canvas.axes2.set_xlim([origin, t])
                    self.MplWidget.canvas.axes2.set_ylim([listy_min, listy_max])
                    self.MplWidget.canvas.axes2.plot(t_list, filtered_y, color='blue')  
                    self.MplWidget.canvas.axes2.set_title('Filt Accelerometer (Y)') 

                    self.MplWidget.canvas.axes3.set_xlabel("Frequency(Hz)")
                    self.MplWidget.canvas.axes3.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes3.set_xlim([0, 200])
                    self.MplWidget.canvas.axes3.set_ylim([0, 1])
                    self.MplWidget.canvas.axes3.plot(x_f, 2.0/N * np.abs(y_f[:N//2]), color='green')  
                    self.MplWidget.canvas.axes3.set_title('FFT Accelerometer (Y)')     

                    self.MplWidget.canvas.figure.tight_layout() #隔開兩個圖
                    self.MplWidget.canvas.draw()
                
            elif self.ckb_Z.isChecked():
                #四分位距法去除離群值
                out_pd = pd.Index(accelerometer_z)
                out_z = remove_outliers(out_pd)

                '''
                nyquist_freq = 0.5 * fs
                low = lowcut / nyquist_freq
                high = highcut / nyquist_freq
                b, a = butter(order, [low, high], btype="bandstop")
                df_x = filtfilt(b, a, out_x)
                '''
                # 設計低通濾波器
                cutoff = highcut  # 截止頻率
                nyq = 0.5 * fs  # Nyquist頻率
                normal_cutoff = cutoff / nyq
                b, a = butter(4, normal_cutoff, btype='low', analog=False)

                # 使用濾波器濾波信號
                filtered_z = filtfilt(b, a, accelerometer_z)

                # 設置參數
                N = sampleRate_rate      # 訊號長度
                T = 1 / sampleRate_rate   # 取樣週期
                y_f = np.fft.fft(filtered_z)
                x_f = np.linspace(0.0, 1.0/(2.0*T), N//2)

                if click_state == True:
                    self.MplWidget.canvas.figure.clf()
                    self.MplWidget.canvas.axes0 = self.MplWidget.canvas.figure.add_subplot(411)
                    self.MplWidget.canvas.axes1 = self.MplWidget.canvas.figure.add_subplot(412)
                    self.MplWidget.canvas.axes2 = self.MplWidget.canvas.figure.add_subplot(413)
                    self.MplWidget.canvas.axes3 = self.MplWidget.canvas.figure.add_subplot(414)

                    self.MplWidget.canvas.axes0.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes0.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes0.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    #self.MplWidget.canvas.axes0.set_xlim([origin, t])
                    self.MplWidget.canvas.axes0.set_ylim([listz_min, listz_max])
                    self.MplWidget.canvas.axes0.plot(t_list, accelerometer_z, color='red')  
                    self.MplWidget.canvas.axes0.set_title('Accelerometer (Z)')

                    self.MplWidget.canvas.axes1.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes1.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes1.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    #self.MplWidget.canvas.axes1.set_xlim([origin, t])
                    self.MplWidget.canvas.axes1.set_ylim([listz_min, listz_max])
                    self.MplWidget.canvas.axes1.plot(t_list, out_z, color='#FF8328')  
                    self.MplWidget.canvas.axes1.set_title('Remove Outliers Accelerometer (Z)')

                    self.MplWidget.canvas.axes2.set_xlabel("Time(sec)")
                    self.MplWidget.canvas.axes2.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes2.set_xlim([origin/sampleRate_rate, t/sampleRate_rate])
                    #self.MplWidget.canvas.axes2.set_xlim([origin, t])
                    self.MplWidget.canvas.axes2.set_ylim([listz_min, listz_max])
                    self.MplWidget.canvas.axes2.plot(t_list, filtered_z, color='blue')  
                    self.MplWidget.canvas.axes2.set_title('Filt Accelerometer (Z)') 

                    self.MplWidget.canvas.axes3.set_xlabel("Frequency(Hz)")
                    self.MplWidget.canvas.axes3.set_ylabel("g(m/s\u00B2)")
                    self.MplWidget.canvas.axes3.set_xlim([0, 200])
                    self.MplWidget.canvas.axes3.set_ylim([0, 1])
                    self.MplWidget.canvas.axes3.plot(x_f, 2.0/N * np.abs(y_f[:N//2]), color='green')  
                    self.MplWidget.canvas.axes3.set_title('FFT Accelerometer (Z)')     

                    self.MplWidget.canvas.figure.tight_layout() #隔開兩個圖
                    self.MplWidget.canvas.draw()
            else:
                self.MplWidget.canvas.figure.clf()
                self.MplWidget.canvas.draw()


    def click_acc(self):
        global click_state
        if self.ckb_acc.isChecked():
            click_state = True
            self.ckb_acc_out.setChecked(False)
            self.ckb_X.setChecked(True)
            self.ckb_Y.setChecked(True)
            self.ckb_Z.setChecked(True)
        else:
            click_state = True
            self.ckb_X.setChecked(False)
            self.ckb_Y.setChecked(False)
            self.ckb_Z.setChecked(False)
    
    def click_acc_out(self):
        global click_state
        global LC
        global HC
        if self.ckb_acc_out.isChecked():
            click_state = True
            self.ckb_acc.setChecked(False)
            self.ckb_X.setChecked(True)
            self.ckb_Y.setChecked(True)
            self.ckb_Z.setChecked(True)
            LC = int(self.line_low.text())
            HC = int(self.line_high.text())
        else:
            click_state = True
            self.ckb_X.setChecked(False)
            self.ckb_Y.setChecked(False)
            self.ckb_Z.setChecked(False)

    def click_X(self):
        global click_state
        global LC
        global HC
        click_state = True
        self.ckb_acc.setChecked(False)
        LC = int(self.line_low.text())
        HC = int(self.line_high.text())

    def click_Y(self):
        global click_state
        global LC
        global HC
        click_state = True
        self.ckb_acc.setChecked(False)
        LC = int(self.line_low.text())
        HC = int(self.line_high.text())

    def click_Z(self):
        global click_state
        global LC
        global HC
        click_state = True
        self.ckb_acc.setChecked(False)
        LC = int(self.line_low.text())
        HC = int(self.line_high.text())
    
    def btn_lock(self):
        global lock
        global time_reset
        if lock == 0:
            lock = 1
            print(lock)
            self.btn_stop.setText('恢復')

        elif lock == 1:
            lock = 0
            print(lock) 
            self.btn_stop.setText('暫停')
    
    def click_record(self):
        global excel_state
        global ex_row
        global wb
        global ws
        global date_time
        self.btn_record.setEnabled(False)
        self.btn_save.setEnabled(True)
        excel_state = True
        ex_row = 1

        date_time = time.strftime("%Y-%m-%d_%H_%M_%S", time.localtime())
        wb = Workbook()
        ws=wb.active
        ws.title = "data"
        '''
        ws.cell(row=ex_row, column=1, value="accelerometer_x")
        ws.cell(row=ex_row, column=2, value="accelerometer_y")
        ws.cell(row=ex_row, column=3, value="accelerometer_z")
        ex_row += 1
        '''
        wb.save('%s.xlsx' % date_time)

    def save_excel(self):
        global ws
        global wb
        global date_time
        global ex_row
        global accelerometer_excel
        global excel_state

        excel_state = False
        self.btn_record.setEnabled(True)
        self.btn_save.setEnabled(False)

        QMessageBox.information(None, '開始存取', '存取中請稍等')

        while len(accelerometer_excel)>0:
            ws.cell(row=ex_row, column=1, value = round(float(accelerometer_excel[0][0]),4))
            ws.cell(row=ex_row, column=2, value = round(float(accelerometer_excel[0][1]),4))
            ws.cell(row=ex_row, column=3, value = round(float(accelerometer_excel[0][2]),4))
            ex_row += 1
            accelerometer_excel.pop(0)
        
        wb.save('%s.xlsx' % date_time)

        QMessageBox.information(None, '存取成功', '存取完畢請確認')

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.groupBox.setTitle(_translate("MainWindow", "主機IP地址"))
        self.line_IP_s.setText(_translate("MainWindow", "0.0.0.0"))
        self.btn_IP.setText(_translate("MainWindow", "查詢"))
        self.groupBox_4.setTitle(_translate("MainWindow", "刀把IP地址"))
        self.line_IP_c.setText(_translate("MainWindow", "0.0.0.0"))
        self.groupBox_2.setTitle(_translate("MainWindow", "顯示設定"))
        self.ckb_acc.setText(_translate("MainWindow", "加速度"))
        self.ckb_acc_out.setText(_translate("MainWindow", "加速度(濾波)"))
        self.ckb_X.setText(_translate("MainWindow", "X"))
        self.ckb_Y.setText(_translate("MainWindow", "Y"))
        self.ckb_Z.setText(_translate("MainWindow", "Z"))
        self.groupBox_3.setTitle(_translate("MainWindow", "加速度:"))
        self.label_5.setText(_translate("MainWindow", "X:"))
        self.label_4.setText(_translate("MainWindow", "Y:"))
        self.label_6.setText(_translate("MainWindow", "Z:"))
        self.groupBox_6.setTitle(_translate("MainWindow", "設置:"))
        self.label_7.setText(_translate("MainWindow", "lowcut:"))
        self.label_8.setText(_translate("MainWindow", "highcut:"))
        self.groupBox_5.setTitle(_translate("MainWindow", "顯示"))
        self.btn_record.setText(_translate("MainWindow", "紀錄"))
        self.btn_save.setText(_translate("MainWindow", "保存"))
        self.btn_stop.setText(_translate("MainWindow", "暫停"))

def remove_outliers(data, threshold1=0.07, threshold2=0.14):
    """
    去除異常值的函式
    :param data: 要處理的資料，為一維 numpy array
    :param threshold: 設定的異常值閥值，預設為 0.1
    :return: 處理後的資料，為一維 numpy array
    """
    result = data.copy()
    for i in range(1, len(data)):
        #diff1 = (abs(data[i]) - abs(data[i-1])) / abs(data[i-1])
        #diff2 = (abs(data[i-1]) - abs(data[i])) / abs(data[i])
        #diff1 = abs(abs(data[i]) - abs(data[i-1]))
        #diff2 = abs(abs(data[i-1]) - abs(data[i]))
        diff = abs(data[i] - data[i-1])
        if (diff >= threshold1) and (diff <= threshold2):
            start = max(0, i-1)
            end = min(len(data), i+2)
            avg = np.mean(data[start:end])
            ret = list(result)
            ret[i] = avg
            result = ret
    return result


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
